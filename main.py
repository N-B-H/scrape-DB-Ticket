#import PyPDF2 as p
#import tabula as t

import os
import pdfplumber
from openpyxl import Workbook
from datetime import datetime
import re


file_names = os.listdir("tickets")
file_paths = [f"tickets\\{file_name}" for file_name in file_names]

#file_name = file_names[0]
#file_path = f"tickets\\{file_name}"


def pdf_to_data(file_path):

    ### pdfplumber ###
    pdf = pdfplumber.open(file_path)
    page = pdf.pages[0] # only first page
    tables = page.extract_tables() #extract tables
    text = page.extract_text()
    last_line = text.splitlines()[-1]

    if tables[0][0][0].startswith('Positionen'): international = True

    str_lines = []

    for table in tables:
        for row in table:
            for cell in row:
                if cell: str_lines.extend(cell.split("\n"))

    ticket_string = '\n'.join(str_lines+[last_line])

    #print(ticket_string)

    ### parse all tickets from ticket_string ###

    # find items with regular expressions
    # dict with attribute as key and re-marker as value
    re_marker = {
        "date": r'(?:(?:Gültigkeit: a.)|(?:Fahrtantritt a.)) (.*)\n', #r'Gültigkeit: a[mb] (.*)\n', #oder Fahrtantritt am 24.10.2020
        "start": 'Hinfahrt: (.*)   ',

        ### BISHER ##
        #"dest": 'Hinfahrt: .*  (.*)[,$]?',

        ### 1. Versuch Tim ###
        #"dest": 'Hinfahrt: .+   ([^,]+),.*$', #r'Hinfahrt: .*  (.*)[,$]?',

        ### 2. Versuch Tim ###
        #"dest": r'Hinfahrt:. + ([ ^,]+)(,.* | $)',

        "dest": r'Hinfahrt: .+   (.+?)[,\n]',

        "cost": 'Betrag (.*)€',
        "ID": r'\n(.+?) Seite 1 / 1$'
    }

    re_marker_international = r'Gütigkeit: a.(.+).+VON ->NACH XX\n.* (.+) ->(.+)'

    travel = {} # will have the same keys as the regular expression markers
    for attribute in re_marker:
        m = re.search(re_marker[attribute], ticket_string)
        try:
            travel[attribute] = m.group(1)
        except:
            print(travel)
            print(attribute)
            print(ticket_string)
            print("\n\n", page.extract_text())
    #print(ticket_string)
    # convert date and cost
    travel["date"] = str(datetime.strptime(travel["date"], '%d.%m.%Y').date())
    travel["cost"] = float(travel["cost"].replace(',', '.'))

    #print(tickets)
    return travel



tickets = [pdf_to_data(file_path) for file_path in file_paths]
for ticket in tickets: print(ticket)

### save to an excel-file
def tickets_to_excel(tickets, xls_file_path):

    wb = Workbook()
    sheet = wb.active

    #initialise the worksheet

    sheet.title = "DB-Reisen"

    for y in range(len(tickets)): # as many rows as there are tickets
        #for x in range(1,5): # 4 attributes
        sheet.cell(row=y+1, column=1).value = tickets[y]["date"]

        sheet.cell(row=y+1, column=2).value = f'{tickets[y]["start"]} - {tickets[y]["dest"]}'

        sheet.cell(row=y+1, column=3).value = tickets[y]["cost"]
        #sheet.cell(row=y, column=3).style = 'Currency'
        sheet.cell(row=y+1, column=3).number_format = '#,##0.00€'


    #adjust cell width:
    '''
    for column_cells in sheet.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        worksheet.column_dimensions[column_cells[0].column].width = length
    '''
        
    wb.save(xls_file_path)


tickets = sorted(tickets, key=lambda x: x['date'], reverse=True)

tickets_to_excel(tickets, 'output.xlsx')

'''for file_path in file_names:
    tickets = pdf_to_data(file_path)'''



'''
tokens = {
    "Gültigkeit: am ": str(), #datetime.date(),
    "Hinfahrt: ": str(),
    "Rückfahrt: ": str(),
    "Betrag " : str() #float()

}

for line in str_lines:
    for token in tokens:
        if token in line:
            if tokens[token] == str():
                tokens[token] = line.replace(token, '')

                if "€" in tokens[token]:
                    i = tokens[token].index("€")
                    tokens[token] = tokens[token][:i]

print(tokens)
'''


### Tables ###
'''
tables = page.extract_tables()
for table in tables:
    print("|-")
    for row in table:
        print("--")
        for cell in row:
            print(cell)
'''




### text (bugged...) ###

#text = pdf.pages[0].extract_text()
#print(text)

### lines (not working as expected) ###
#lines = page.lines
#print(lines)

### words (not working as expected) ###
'''
words = page.extract_words()
print(words)
'''
# open all files and read to list
#all_files = [open(f"tickets\\{file_names[i]}", "rb") for i in range(len(file_names))]
#all_files = [t.read_pdf(f"tickets\\{file_names[0]}", pages=1, multiple_tables=True) for i in range(len(file_names))]
#file = all_files[0]


#output_file = open(file_names[0].replace(".pdf",".txt"), "w+")
#output_file = open("output.csv", "w+")

#t.convert_into("example.pdf", "output.csv")




